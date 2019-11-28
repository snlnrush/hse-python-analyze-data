"""
Вам дан xlsx-файл с информацией о спортивных площадках в Москве.
Вам необходимо сдать на проверку json-файл, в котором будет храниться один словарь,
ключами которого будут административные округа (AdmArea), а значениями словари, в которых,
в свою очередь, ключами будут названия районов (District), относящихся к этому административному округу,
а значениями - списки адресов площадок (Address) в том порядке, в котором они встречались в исходном файле.

Ваш файл должен выглядеть примерно так:

{"Северо-Западный административный округ": {"район Строгино": ["улица Исаковского, дом 24, корпус 1",
"Неманский проезд, дом 9"], "район Северное Тушино": ["улица Свободы, дом 56", "улица Свободы, дом 56",
"улица Свободы, дом 56", "улица Свободы, дом 56"], "район Покровское-Стрешнево": ["Иваньковское шоссе, дом 6"],
"район Щукино": ["Сосновая улица, дом 3, строение 2"]}, ...}
"""
import openpyxl
import json


wb = openpyxl.load_workbook('data-25290-2019-10-30.xlsx')
ws = wb.active
main_dict = {}

for val in range(2, ws.max_row + 1):
    adm_area = ws.cell(row=val, column=5).value
    district = ws.cell(row=val, column=6).value
    address = ws.cell(row=val, column=7).value
    if adm_area not in main_dict:
        main_dict[adm_area] = {district: [address, ]}
    elif district not in main_dict[adm_area]:
        main_dict[adm_area].update({district: [address, ]})
    else:
        main_dict[adm_area][district].append(address)

fout = open('sport_places.json', 'w', encoding='utf-8')
json.dump(main_dict, fout)
fout.close()
