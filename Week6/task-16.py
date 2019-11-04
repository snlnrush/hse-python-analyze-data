"""
Васю назначили завхозом в туристической группе и он подошёл к подготовке ответственно,
составив справочник продуктов с указанием калорийности на 100 грамм, а также содержание белков,
жиров и углеводов на 100 грамм продукта. Ему не удалось найти всю информацию,
поэтому некоторые ячейки остались незаполненными (можно считать их значение равным нулю).
Также он использовал какой-то странный офисный пакет и разделял целую и дробную часть чисел запятой.
Таблица доступна по ссылке https://stepik.org/media/attachments/lesson/245290/trekking2.xlsx

Вася составил раскладку по продуктам на один день (она на листе "Раскладка") с указанием названия продукта
 и его количества в граммах. Посчитайте 4 числа: суммарную калорийность и граммы белков, жиров и углеводов.
 Числа округлите до целых вниз и введите через пробел.
"""
import openpyxl

wb = openpyxl.load_workbook('trekking2-245290.xlsx')
ws1 = wb['Справочник']
ws2 = wb['Раскладка']
dict_catalog = {}
dict_layout = {}
lst_sum = []

for y in range(ws2.min_column + 1, ws2.max_row + 1):
    product = ws2.cell(row=y, column=1).value
    weight = ws2.cell(row=y, column=2).value
    dict_layout[product] = dict_layout.get(product, 0) + weight

for y in range(ws1.min_column + 1, ws1.max_row + 1):
    product = ws1.cell(row=y, column=1).value
    cal = ws1.cell(row=y, column=2).value
    if ws1.cell(row=y, column=3).value is None:
        squirrels = 0
    else:
        squirrels = ws1.cell(row=y, column=3).value
    if ws1.cell(row=y, column=4).value is None:
        fats = 0
    else:
        fats = ws1.cell(row=y, column=4).value
    if ws1.cell(row=y, column=5).value is None:
        carbohydrates = 0
    else:
        carbohydrates = ws1.cell(row=y, column=5).value
    dict_catalog[product] = cal, squirrels, fats, carbohydrates

for key, val in dict_layout.items():
    perc = val / 100
    lst_sum.append((dict_catalog[key][0] * perc, dict_catalog[key][1] * perc, dict_catalog[key][2] * perc, dict_catalog[key][3] * perc))

sum_cal = 0
sum_sq = 0
sum_fats = 0
sum_carb = 0

for item in lst_sum:
    sum_cal += item[0]
    sum_sq += item[1]
    sum_fats += item[2]
    sum_carb += item[3]

print(int(sum_cal), int(sum_sq), int(sum_fats), int(sum_carb))
