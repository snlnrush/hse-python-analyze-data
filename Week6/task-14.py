"""
Вася планирует карьеру и переезд. Для это составил таблицу,
в которой для каждого региона записал зарплаты для разных интересные ему профессий.
Таблица доступна по ссылке https://stepik.org/media/attachments/lesson/245267/salaries.xlsx.
Выведите название региона с самой высокой медианной зарплатой (медианой называется элемент,
стоящий в середине массива после его упорядочивания) и, через пробел,
название профессии с самой высокой средней зарплатой по всем регионам.
"""
import openpyxl

lst_salary = []

wb = openpyxl.load_workbook('salaries-245267.xlsx')
ws = wb.active
d1t = {}
lst_t = []
df = {}
df_average = {}
lst_average = []
lst_av_f = []

for y in range(2, 10):
    cities = ws.cell(row=y, column=1).value
    for x in range(2, 9):
        profession = ws.cell(row=1, column=x).value
        salary = ws.cell(row=y, column=x).value
        d1t[salary] = cities, profession
        lst_t.append(int(salary))
        for j in range(2, 10):
            prof_salary = ws.cell(row=j, column=x).value
            lst_average.append(prof_salary)
        ave_sal = int(int(sum(lst_average)) / len(lst_average))
        df_average[str(ave_sal)] = profession
        lst_av_f.append(ave_sal)

    lst_t.sort()
    median = lst_t[len(lst_t) // 2]
    df[median] = d1t[median]
    d1t.clear()
    lst_t.clear()

lst_t = list(df.keys())
lst_t.sort()
print(df[lst_t[-1]][0], df_average[max(df_average.keys())])
