"""
Главный бухгалтер компании "Рога и копыта" случайно удалил ведомость с начисленной зарплатой.
К счастью, у него сохранились расчётные листки всех сотрудников.
Помогите по этим расчётным листкам восстановить зарплатную ведомость.
Архив с расчётными листками доступен по ссылке https://stepik.org/media/attachments/lesson/245299/rogaikopyta.zip
(вы можете скачать и распаковать его вручную или самостоятельно научиться делать это с помощью скрипта на Питоне).

Ведомость должна содержать 1000 строк, в каждой строке должно быть указано ФИО сотрудника и,
через пробел, его зарплата. Сотрудники должны быть упорядочены по алфавиту.
"""
import openpyxl

lst_salary = []

for i in range(1, 1001):
    wb = openpyxl.load_workbook('{}.xlsx'.format(str(i)))
    ws = wb.active
    lst_salary.append((ws['B2'].value, ws['D2'].value))

lst_salary.sort()

fout = open('out-text.txt', 'w', encoding='utf8')

for fio, salary in lst_salary:
    print(fio, salary, file=fout)

fout.close()
